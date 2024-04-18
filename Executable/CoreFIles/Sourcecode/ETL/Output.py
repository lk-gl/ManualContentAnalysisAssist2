import openpyxl
import ETL.Files as Files



def initTables(TargetDir):
    path = TargetDir + "\\Results.xlsx"
    wb = openpyxl.load_workbook(path)
    for obj in Tables.tableLst:
        ws = wb[obj.sheet]
        temp_col = obj.rBottom[0]
        if "Link" not in obj.rf: #"Link" doesnt require Table (is one cell)
            if obj.rBottom[1] != obj.lTop[1]:
                obj.lTop = ws.cell(row = obj.lTop[1], column = obj.lTop[0]).coordinate
                obj.rBottom = ws.cell(row = obj.rBottom[1], column = obj.rBottom[0]).coordinate
                obj.applystyle(ws)
                tab = obj.tabellarize()
                tab = tab
                ws.add_table(tab)
            if obj.sheet == "AllData": #Add Files links to raw data sheet
                obj.linkify(ws, obj.rBottom[0])
        if "Link" in obj.rf: #Set link within all other sheets
                cell = ws.cell(row = 2, column = temp_col)
                obj.linkify(ws, cell)  
    wb.save(path)

##############Class##################
class Tables: 
    tableLst = []
    def __init__(self, rf, sheet):
        self.lTop = None
        self.rf = rf
        self.sheet = sheet
        self.rBottom = None
        self.tb = None
        Tables.tableLst.append(self)

    def scope(self, offsetC, offsetR, df):
        self.lTop = (offsetC + 1, offsetR + 1)
        self.rBottom = (offsetC + len(df.columns) + 1, offsetR + len(df.index) + 1)

    def tabellarize(self):
        tab = openpyxl.worksheet.table.Table(displayName=self.rf, ref= str(self.lTop +":"+self.rBottom))
        self.tb = tab
        return tab
    
    def linkify(self, ws, col):
        temp = Files.TargetFiles.OutletSt
        temp.add("AllCodedFiles")
        if ws.title in temp:
            if type(col.value) == str: 
                col.hyperlink = col.value; col.font = openpyxl.styles.Font(color="0000EE")
                col.value = "OpenFile"   
        elif ws.title == "AllData":
            for section in ws[self.lTop:self.rBottom]:
                for cell in section:
                    if type(cell.value)== str:
                        if cell.value.endswith(".docx"):
                            cell.hyperlink = cell.value; cell.font = openpyxl.styles.Font(color="0000EE")
                            cell.value = "OpenFile"
    def applystyle(self, ws):
        for section in ws[self.lTop:self.rBottom]:
            for cell in section:
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                cnt = 0
                if cell.border.top.style  is not None: cnt += 1
                if cell.border.left.style  is not None: cnt += 1
                if cell.border.bottom.style  is not None: cnt += 1
                if cell.border.right.style  is not None: cnt += 1
                if cnt == 4: 
                    cell.fill = openpyxl.styles.PatternFill(start_color="D9D9D9", fill_type = "solid")
                    if cell.value in ["Positivity","Share","Entries/Article"]:
                        r = cell.row 
                        while r <= ws[self.rBottom].row: 
                            if cell.value == "Entries/Article":
                                ws[cell.column_letter + str(r)].number_format = '0.00'
                            else:
                                ws[cell.column_letter + str(r)].number_format = '0.0%'
                            r +=1 
                        



def printattr(lst):
        for obj in lst:
            print("Left:", obj.lTop, "Right:", obj.rBottom, "REf:",obj.rf)
