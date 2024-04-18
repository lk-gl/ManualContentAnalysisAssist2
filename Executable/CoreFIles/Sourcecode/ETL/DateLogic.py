# Needed: Time Harmonizer and Transform to Object
# Later: Run Another Analysis step which creates the time based analyisis 

#from dateutil import parser as dateparser

#5*Y

from datetime import datetime
import pandas as pd
import ETL.Files as Files
import ETL.Entries as Entries 
import ETL.ToExcel as ToExcel
import ETL.Categories as Categories
import re
import openpyxl

def harmonize(date): 
    dateObj = datetime.strptime(date, '%Y-%m-%d').date()
    return dateObj

def initDateSlices(Interval):
    mult = int(Interval.split("<x>")[0])
    base = Interval.split("<x>")[1]
    periods = pd.to_datetime(max(Files.TargetFiles.DateSt)).to_period(freq=base) - pd.to_datetime(min(Files.TargetFiles.DateSt)).to_period(freq=base)
    p = periods.freqstr
    periods_int = int(re.findall(r'\d+', p)[0]) 
    periods_int = periods_int // mult + (periods_int % mult > 0) + 1
    dct = {}; lst =[]
    for entry in Entries.CodedEntries.CodedEntryLst: 
        dc = {"Time":pd.Timestamp(entry.date), "FileNr":entry.fileEnum, "Category":entry.cat, "Polarity":entry.PN,"Outlet":entry.outlet,"ix":entry.index}
        u = pd.DataFrame.from_dict(data=[dc], orient="columns")
        lst.append(u)
    df = pd.concat(lst,axis=0)
    
    df.index = df["ix"]
    df = df.sort_index()
    bins_dt = pd.date_range(min(Files.TargetFiles.DateSt)+ pd.offsets.YearBegin(-1)+pd.offsets.MonthBegin(-1), freq=str(mult)+base, periods=periods_int)
    if base == "Q":
        for i, bin in enumerate(bins_dt):
            if bin + pd.offsets.MonthBegin(4*mult) < min(Files.TargetFiles.DateSt):
                bins_dt = bins_dt.delete(i) 
    elif base == "M":
        for i, bin in enumerate(bins_dt):
            if bin + pd.offsets.MonthBegin(1*mult) < min(Files.TargetFiles.DateSt):
                bins_dt = bins_dt.delete(i) 
    bins_str = bins_dt.astype(str).values
    labels = ['{}to{}'.format(bins_str[i-1], bins_str[i]) for i in range(1, len(bins_str))]
    df["Groups"]= pd.cut(df["Time"], bins=bins_dt, labels=labels, include_lowest=True)
    for group in labels:    
        df2 = df[df["Groups"] == group]
        if pd.Timestamp(group.split("to")[1]) > pd.Timestamp(max(Files.TargetFiles.DateSt)):
            group = group.split("to")[0] + "to" + str(max(Files.TargetFiles.DateSt))
        df2.Name = group
        for file in Files.TargetFiles.FileLst: 
            if file.nr in list(df2["FileNr"].unique()):
                file.group = group
        DateSlices(group, df2)
        ####From here untested###### 
        if df2.empty==False:
            ToExcel.AnalysisFrames(df2, "DATE" + group, "", "", "", "", "").numeralize(df2)
        if df2.empty==True:
            ToExcel.AnalysisFrames(df2, "DATE" + group, "", "", "", "", "").numeralize(df2)


def buildChart(TargetDir, base):
    path = TargetDir + "\\Results.xlsx"
    wb = openpyxl.load_workbook(path)
# Plan: - 1 Chart per Cat, displaying development of Share and Positivity
#       - 1 Chart which shows development of all Cats
#       - 1 Chart which shows start to finish development of Positivity of Cats
    wb.create_sheet("Charts")
    ws = wb["Charts"]

    Share = GetChartData(wb["ByDateSplice"],6,3, base)
    CategoryTotal = GetChartData(wb["ByDateSplice"],6,4, base)
    CategoryPositive = GetChartData(wb["ByDateSplice"],6,5, base)
    CategoryNegativ = GetChartData(wb["ByDateSplice"],6,6, base)
    CategoryPositivity = GetChartData(wb["ByDateSplice"],6,7, base)
    chartcount = 0
    initchart(ws,Share,chartcount, "Share of Categories")
    chartcount+=len(DateSlices.SlicesLst)+1
    initchart(ws,CategoryTotal,chartcount, "Total Category Frequency")
    chartcount+=len(DateSlices.SlicesLst)+1
    initchart(ws,CategoryPositive,chartcount, "Cat Positives")
    chartcount+=len(DateSlices.SlicesLst)+1
    initchart(ws,CategoryNegativ,chartcount, "Cat Negatives")
    chartcount+=len(DateSlices.SlicesLst)+1
    initchart(ws,CategoryPositivity,chartcount, "Cat Positivity")
    chartcount+=len(DateSlices.SlicesLst)+1
    wb.save(TargetDir+"\\Results.xlsx")

def initchart(ws,dta,chartcount,tlt):
    chartrow = chartcount
    for row in dta:
        ws.append(row)
    c1 = openpyxl.chart.LineChart()
    c1.title = tlt + " by DateSplice"
    c1.style = 2
    c1.y_axis.title = "Prominence"
    c1.x_axis.title = 'Interval'
    xcats = openpyxl.chart.Reference(ws, min_col=1, min_row=chartrow+2, max_col=1, max_row=len(DateSlices.SlicesLst)+chartrow+1)
    dta = openpyxl.chart.Reference(ws, min_col=2, min_row=1+chartrow, max_col=11, max_row=len(DateSlices.SlicesLst)+1+chartrow)        
    c1.add_data(dta, titles_from_data=True)
    c1.set_categories(xcats)
    ws.add_chart(c1, ws.cell(row = 1+chartrow, column = 10).coordinate)

def GetChartData(ws,startrow, startcol, base): 
    row = startrow
    col = startcol
    bs = ["Interval", Categories.CatGlobal.CatLst[0].name,Categories.CatGlobal.CatLst[1].name,Categories.CatGlobal.CatLst[2].name,Categories.CatGlobal.CatLst[3].name,
            Categories.CatGlobal.CatLst[4].name,Categories.CatGlobal.CatLst[5].name,Categories.CatGlobal.CatLst[6].name,Categories.CatGlobal.CatLst[7].name,Categories.CatGlobal.CatLst[8].name, Categories.CatGlobal.CatLst[9].name]
    chart_lst =[bs]
    
    for slice in DateSlices.SlicesLst:
        grouptag = []
        arr = slice.group.split("to")
        for date in arr: 
            arr2 = date.split("-")
            if base == "M":
                grouptag.append(arr2[1]+"."+arr2[0][2:])
            if base == "Y":
                grouptag.append(arr2[0][2:])
            if base == "Q":
                if int(arr2[1]) <= 3:
                    grouptag.append("Q1 "+str(arr2[0][2:]))
                elif int(arr2[1]) <= 6:
                    grouptag.append("Q2 "+str(arr2[0][2:]))
                elif int(arr2[1]) <= 9:
                    grouptag.append("Q3 "+str(arr2[0][2:]))
                elif int(arr2[1]) <= 12:
                    grouptag.append("Q4 "+str(arr2[0][2:]))
        grouptag = "-".join(grouptag)
        clst = [grouptag]
        for cat in range (0,10):
            clst.append(ws.cell(row=row,column=col).value)
            row = row +1
        chart_lst.append(clst)
        row = startrow
        col +=8
    return chart_lst
    



class DateSlices:
   SlicesLst = []
   def __init__(self, group, data):
       self.group = group
       self.files = list(data["FileNr"].unique())
       self.data = data
       self.catfreq = None
       DateSlices.SlicesLst.append(self)
       self.pos = None
       self.neg = None
       


# Zieloutput: 
    # - New Sheet: Per Unit Analysis, mirroring Overview, i.e. analysis per outlet and category
    # - Overview: Graphs which show development per Unit
        # By Outlet: Coded Articles and Arguments
        # By Cat (Share)
        # By Cat (Absolutes)
        # By Positivity (P/N)
        # Show data
        # Maybe shit via Pandas-Pivot? 


    

